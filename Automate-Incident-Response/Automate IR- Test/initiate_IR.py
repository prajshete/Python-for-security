from elasticsearch import Elasticsearch
from elasticsearch.exceptions import ConnectionError, TransportError, NotFoundError
from requests.auth import HTTPBasicAuth
from openpyxl.styles import PatternFill
import pandas as pd
import urllib3,requests, ssl, vt, os, asyncio, time, typer, json, smtplib, base64, openpyxl, mimetypes, zipfile, subprocess, shutil
from email.message import EmailMessage
from email.utils import formataddr
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from dotenv import load_dotenv

 
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = typer.Typer()


#=========================== Read API KEY ===========================

def read_api_key(api_path):
    try:
        with open(api_path, "r") as file:
            api_key = file.read().strip()  # Read API key and strip any surrounding whitespace/newlines
        return api_key
    except FileNotFoundError:
        print(f"Error: File '{api_path}' not found.")
        return None
    except Exception as e:
        print(f"Error reading API key: {e}")
        return None

#=======================================================================

#=========================== Read API KEY ===========================

def get_vt_api_key(api_path):
    try:
        with open(api_path, "r") as file:
            api_key = file.read().strip()  # Read API key and strip any surrounding whitespace/newlines
        return api_key
    except FileNotFoundError:
        print(f"Error: File '{api_path}' not found.")
        return None
    except Exception as e:
        print(f"Error reading API key: {e}")
        return None

#=======================================================================




#=============== Function to read FP domains from excel file =========================

def read_false_positive_domains(excel_path, sheet_name='Sheet1'):
    # Read the Excel file
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    # Assuming the first column contains the false positive domains
    domains = df.iloc[:, 0].tolist()
    return domains

#=====================================================================================


#=============== Function to read FP domains from excel file =========================

def build_elasticsearch_query_dns(domains):
    # Build the query string
    query_string_dns = " OR ".join([f"dns.question.name:{domain}" for domain in domains])
    #query_string_dns = f'({query_string_dns})'
    return query_string_dns

#=====================================================================================


#=============== Function to read FP domains from excel file =========================

def build_elasticsearch_query_http(domains):
    # Build the query string
    query_string_http = " OR ".join([f"http.request.body.content:{domain}" for domain in domains])
    return query_string_http

#=====================================================================================

 

# ========================== GET LOOKUP_RESULT LOGS FROM ELASTIC SEARCH ========================

#@app.command()
def get_dns_network_logs(index_name, timestamp: str, fp_domain_query_dns, API_KEY):
    # Connect to Elasticsearch
    try:


        # Initialize the Elasticsearch client
        client = Elasticsearch("https://chromewell-elk-soc.es.us-central1.gcp.cloud.es.io",
        api_key=API_KEY
        )

        if timestamp.endswith('h'):
            hours = int(timestamp[:-1])
     
        index_name = ".ds-logs-endpoint.events.network-default*"
        # Define the query with size inside the body
        query = f'''
        {{
            "size": 1000,
            "query": {{
                "bool": {{
                    "must": [
                        {{"match": {{"event.dataset": "endpoint.events.network"}}}},
                        {{"match": {{"event.action": "lookup_result"}}}},
                        {{"terms": {{"process.name": ["chrome.exe", "OUTLOOK.EXE"]}}}},
                        {{"range": {{"@timestamp": {{"gte": "now-{hours}h", "lte": "now/h"}}}}}}
                    ],
                    "must_not": [
                        {{"query_string": {{"query": "{fp_domain_query_dns}"}}}},
                        {{"terms": {{"user.name": ["SYSTEM", "NETWORK SERVICE", "USER", "adity", "Dell"]}}}},
                        {{
                            "range": {{
                                "destination.ip": {{
                                    "gte": "192.168.0.0",
                                    "lte": "192.168.255.255"
                                }}
                            }}
                        }}
                    ]
                }}
            }}
        }}
        '''
       

        # Fetch logs from Elasticsearch
        response = client.search(index=index_name, body=query)

        # Extract destination IPs from logs
        logs = []
        for hit in response['hits']['hits']:
            source = hit['_source']
            logs.append({
                'user.name': source.get('user', {}).get('name', None),
                'process.name': source.get('process', {}).get('name', None),
                'host.name': source.get('host', {}).get('name', None),
                'event.action': source.get('event', {}).get('action', None),
                'dns.question.name': source.get('dns', {}).get('question', {}).get('name', None)
            })

        return logs

    except ConnectionError as e:
        print(f"ConnectionError connecting to Elasticsearch: {e}")
    except TransportError as e:
        print(f"TransportError connecting to Elasticsearch: {e}")
    except NotFoundError as e:
        print(f"NotFoundError: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    return []
#==============================================================================================



# ========================== GET HTTP_REQUEST LOGS FROM ELASTIC SEARCH ========================

#@app.command()
def get_http_network_logs(index_name, timestamp: str, fp_domain_query_http, API_KEY):
    # Connect to Elasticsearch
    try:
        

        # Initialize the Elasticsearch client
        client = Elasticsearch("https://chromewell-elk-soc.es.us-central1.gcp.cloud.es.io",
        api_key=API_KEY
        )

        if timestamp.endswith('h'):
            hours = int(timestamp[:-1])

        index_name = ".ds-logs-endpoint.events.network-default*"
        # Define the query with size inside the body
        query = f'''
        {{
            "size": 1000,
            "query": {{
                "bool": {{
                    "must": [
                        {{"match": {{"event.dataset": "endpoint.events.network"}}}},
                        {{"match": {{"event.action": "http_request"}}}},
                        {{"terms": {{"process.name": ["chrome.exe", "OUTLOOK.EXE"]}}}},
                        {{"range": {{"@timestamp": {{"gte": "now-{hours}h", "lte": "now/h"}}}}}}
                    ],
                    "must_not": [
                        {{"match": {{"source.ip": "127.0.0.1"}}}},
                        {{"terms": {{"destination.ip": ["13.127.228.11", "40.100.0.0/16", "52.98.88.72", "142.250.0.0/16", "142.251.0.0/16", "192.168.0.0/16", "172.18.0.0/16", "127.0.0.1", "13.127.228.11", "172.31.0.0/16", "210.16.92.198", "103.36.71.187", "23.217.0.0/15"
                     ]}}}},
                        {{"query_string": {{"query": "{fp_domain_query_http}"}}}},
                        {{"terms": {{"user.name": ["SYSTEM", "NETWORK SERVICE", "USER", "adity", "Dell"]}}}},
                        {{
                            "range": {{
                                "destination.ip": {{
                                    "gte": "192.168.0.0",
                                    "lte": "192.168.255.255"
                                }}
                            }}
                        }}
                    ]
                }}
            }}
        }}
        '''
        
       

        # Fetch logs from Elasticsearch
        response = client.search(index=index_name, body=query)

        # Extract destination IPs from logs
        logs = []
        for hit in response['hits']['hits']:
            source = hit['_source']
            logs.append({
                'user.name': source.get('user', {}).get('name', None),
                'process.name': source.get('process', {}).get('name', None),
                'host.name': source.get('host', {}).get('name', None),
                'event.action': source.get('event', {}).get('action', None),
                'source.ip': source.get('source', {}).get('ip', None),
                'destination.ip': source.get('destination', {}).get('ip', None),
                'http.request.body.content': source.get('http', {}).get('request', {}).get('body', {}).get('content', None)
            })

        return logs

    except ConnectionError as e:
        print(f"ConnectionError connecting to Elasticsearch: {e}")
    except TransportError as e:
        print(f"TransportError connecting to Elasticsearch: {e}")
    except NotFoundError as e:
        print(f"NotFoundError: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    return []

#==============================================================================================


#================= CHECKS IP ADDRESSES FOR HITS ON VIRUSTOTALS ===================


def check_ip_virustotal(ip, vt_client):
    try:
        vt_result = vt_client.get_object(f"/ip_addresses/{ip}")
        # Extracting community score from the result
        if hasattr(vt_result, 'last_analysis_stats'):
            community_score = vt_result.last_analysis_stats['malicious']
            return community_score
        else:
            return None
    except vt.APIError as e:
        print(f"VirusTotal API Error: {e}")
    return None

#=================================================================================


#================= CHECKS DNS FOR HITS ON VIRUSTOTALS ============================

def check_dns_virustotal(dns, vt_client):
    try:
        vt_result = vt_client.get_object(f"/domains/{dns}")
        if hasattr(vt_result, 'last_analysis_stats'):
            community_score = vt_result.last_analysis_stats['malicious']
            return community_score
        else:
            return None
    except vt.APIError as e:
        print(f"VirusTotal API Error: {e}")
    return None

#=================================================================================


#==================== EXTRACTS DNS FROM FIELD ====================================

def extract_dns(http_request_body_content):
    if http_request_body_content and "Host:" in http_request_body_content:
        return http_request_body_content.split("Host:")[1].strip().split()[0]
    return None

#=================================================================================


# ================== Iterate through unique destination IPs and check with VirusTotal ==========================


def process_logs_http(vt_client, df):
    confirmed_mal_results = []
    sus_mal_results = []
    clean_results = []
    final_results_http = []


#------------------------------- Call Vt and assign score ---------------------------------------


    for ip in df['destination.ip'].unique():
        community_score = check_ip_virustotal(ip, vt_client)
        if community_score is not None:
            user_name = df.loc[df['destination.ip'] == ip, 'user.name'].iloc[0]
            host_name = df.loc[df['destination.ip'] == ip, 'host.name'].iloc[0]
            dns = df.loc[df['destination.ip'] == ip, 'dns'].iloc[0]
            if community_score >= 5:
                confirmed_mal_results.append({
                    'IP': ip,
                    'Community Score': community_score,
                    'User Name': user_name,
                    'Host Name': host_name,
                    'Status': 'Confirmed malicious'
                })
            elif 1 < community_score <= 5:
                
                dns_score = check_dns_virustotal(dns, vt_client)
                if dns_score is not None:
                    confirmed_mal_results.append({
                        'IP': ip,
                        'Community Score': community_score,
                        #'DNS': dns,
                        #'DNS Score': dns_score,
                        'User Name': user_name,
                        'Host Name': host_name,
                        'Status': 'Confirmed malicious'
                    })
                else:
                    sus_mal_results.append({
                        'IP': ip,
                        'Community Score': community_score,
                        'User Name': user_name,
                        'Host Name': host_name
                    })
            elif community_score == 0:
                dns_score = check_dns_virustotal(dns, vt_client)
                if dns_score != 0:
                    sus_mal_results.append({
                        'IP': ip,
                        'Community Score': community_score,
                        #'DNS': dns,
                        #'DNS Score': dns_score,
                        'User Name': user_name,
                        'Host Name': host_name
                    })
                else:
                    clean_results.append({
                        'IP': ip,
                        'Community Score': community_score,
                        'User Name': user_name,
                        'Host Name': host_name
                    })

        else:
            print(f"Could not get VirusTotal community score for IP: {ip}")

    
    final_results_http={
        'malicious_http': confirmed_mal_results,
        'suspicious_http': sus_mal_results,
        'clean_http': clean_results
    }

    return final_results_http

#======================================================================================================


# ================== Iterate through unique DNS and check with VirusTotal ==========================


def process_logs_dns(vt_client, df):
    confirmed_mal_results = []
    sus_mal_results = []
    clean_results = []
    final_results_dns=[]


#------------------------------- Call Vt and assign score ---------------------------------------

    for dns in df['dns.question.name'].unique():
        community_score = check_dns_virustotal(dns, vt_client)
        if community_score is not None:
            user_name = df.loc[df['dns.question.name'] == dns, 'user.name'].iloc[0]
            host_name = df.loc[df['dns.question.name'] == dns, 'host.name'].iloc[0]
            dns = df.loc[df['dns.question.name'] == dns, 'dns.question.name'].iloc[0]
            if community_score >= 5:    
                confirmed_mal_results.append({
                    'DNS': dns,
                    'DNS Community Score': community_score,
                    'User Name': user_name,
                    'Host Name': host_name,
                    'Status': 'Confirmed malicious'
                })
            elif 1 < community_score < 5:
                sus_mal_results.append({
                    'DNS': dns,
                    'DNS Community Score': community_score,
                    'User Name': user_name,
                    'Host Name': host_name
                })
            elif community_score == 0:
                clean_results.append({
                    'DNS': dns,
                    'DNS Community Score': community_score,
                })
        else:
            print(f"Could not get VirusTotal community score for DNS: {dns}")


# ------------------------merge and return results ------------------------


    final_results_dns={
        'malicious_dns': confirmed_mal_results,
        'suspicious_dns': sus_mal_results,
        'clean_dns': clean_results
    }

    return final_results_dns

#======================================================================================================


#=========================== SEND RESPONSE EMAIL ================================

def send_email(final_results, in_dir):

    print("\n[+] Sending Alert Email . . .")
    port = 587  # For starttls
    smtp_server = "smtp.gmail.com"
    sender_email = "prajshete17@gmail.com"
    receiver_email = "praj@pkfalgosmic.com"
    p = "bHBjciBvamdkIGdnemogc29oaQ=="
    p = base64.b64decode(p) 
    p = p.decode("ascii")
    file_names = ['output_http.xlsx', 'output_dns.xlsx']
    timeout=120

    
    ip_list = []
    dns_list = []
    ip_list_defang = []
    dns_list_defang = []
   
    if 'http' in final_results and 'malicious_http' in final_results['http']:
        for IP in final_results['http']['malicious_http']: 
            ip_list.append(IP['IP'])

    for ip in ip_list:
        defanged_ip = ip.replace('.','[.]')
        ip_list_defang.append(defanged_ip)
    
    ip_list_str = "\n".join(ip_list_defang)
    
   
    for dns in final_results['dns']['malicious_dns']: 
        dns_list.append(dns['DNS'])
    
    for domain in dns_list:
        defanged_dns = domain.replace('.','[.]')
        dns_list_defang.append(defanged_dns)

    dns_list_str = "\n".join(dns_list_defang)
    #print(dns_list_str)
    
    
    body = f"""\

    Network Connections have been detected to the following external malicious IP/DNS addresses:

    IP's:
    {ip_list_str}

    DNS's:
    {dns_list_str}
    
    [+] Compromised Hosts have been Isolated

    Review the attached log files for each user in the ZIP for more details!
    """

    # Create the email message
    message = MIMEMultipart()
    message["From"] = formataddr(("Elastic Alert", sender_email))
    message["To"] = receiver_email
    message["Subject"] = "!! ALERT - Malicious Network Connections !!"
  
    # Add body to email
    message.attach(MIMEText(body, "plain"))
    
    #Attach Log Files
    files = [f for f in os.listdir(in_dir) if os.path.isfile(os.path.join(in_dir, f))]

    for file_name in file_names:
        file_path = os.path.join(in_dir, file_name)
        if os.path.exists(file_path):
            with open(file_path, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename= {os.path.basename(file_name)}")
                message.attach(part)
        else:
            print(f"File not found: {file_path}")

    
    today = datetime.now().strftime("%m-%d")
    dir_to_zip = f"C:/Users/prajs_28/OneDrive/Desktop/PKF ALgosmic/Automate IR/Automate-Incident-Response/Compromised_logs_{today}"
    zip_attachment = f"C:/Users/prajs_28/OneDrive/Desktop/PKF ALgosmic/Automate IR/Automate-Incident-Response/Compromised_logs_{today}"
    shutil.make_archive(zip_attachment, 'zip', dir_to_zip)
    zip_file = f"{zip_attachment}.zip"
    # Attach zip file to email
    if os.path.exists(zip_file):
        with open(zip_file, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename=Compromised_logs_{today}.zip")
            message.attach(part)
    else:
        print(f"Zip file not found: {zip_attachment}")


    #Sending Email
    context = ssl.create_default_context()
    with smtplib.SMTP(smtp_server, port, timeout) as server:
        server.starttls(context=context)
        server.login(sender_email, p)
        server.sendmail(sender_email, receiver_email, message.as_string())
        server.quit()

    print("[+] Email Successfully Sent !!")

#=======================================================================================================



#=========================== get compromised http logs ==============================
def get_compromised_http_dns_logs(user, fp_domain_query_http, fp_domain_query_dns, timestamp, API_KEY):
    # Connect to Elasticsearch
    try:
        

        # Initialize the Elasticsearch client
        client = Elasticsearch("https://chromewell-elk-soc.es.us-central1.gcp.cloud.es.io",
        api_key=API_KEY
        )

        
        index_name = ".ds-logs-endpoint.events.network-default*"
        # Define the query with size inside the body
        
        query = f'''
        {{
            "size": 1000,
            "query": {{
                "bool": {{
                    "must": [
                        {{"match": {{"event.dataset": "endpoint.events.network"}}}},
                        {{"terms": {{"event.action": ["http_request", "lookup_result"]}}}},
                        {{"terms": {{"process.name": ["chrome.exe", "OUTLOOK.EXE"]}}}},
                        {{"range": {{"@timestamp": {{"gte": "now-{timestamp}h", "lte": "now/h"}}}}}},
                        {{"match": {{"user.name":"{user}"}}}}
                    ],
                    "must_not": [
                        {{"match": {{"source.ip": "127.0.0.1"}}}},
                        {{"terms": {{"destination.ip": ["13.127.228.11", "40.100.0.0/16", "52.98.88.72", "142.250.0.0/16", "142.251.0.0/16", "192.168.0.0/16", "172.18.0.0/16", "127.0.0.1", "13.127.228.11", "172.31.0.0/16", "210.16.92.198", "103.36.71.187", "23.217.0.0/15"
                     ]}}}},
                        {{"query_string": {{"query": "{fp_domain_query_http}"}}}},
                        {{"query_string": {{"query": "{fp_domain_query_dns}"}}}},
                        {{"terms": {{"user.name": ["SYSTEM", "NETWORK SERVICE", "USER", "adity", "Dell"]}}}},
                        {{
                            "range": {{
                                "destination.ip": {{
                                    "gte": "192.168.0.0",
                                    "lte": "192.168.255.255"
                                }}
                            }}
                        }}
                    ]
                }}
            }}
        }}
        '''
        
        

        # Fetch logs from Elasticsearch
        response = client.search(index=index_name, body=query)

        # Extract destination IPs from logs
        logs = []
        for hit in response['hits']['hits']:
            source = hit['_source']
            logs.append({
                '@timestamp': source.get('@timestamp', None),
                'user.name': source.get('user', {}).get('name', None),
                'process.name': source.get('process', {}).get('name', None),
                'host.name': source.get('host', {}).get('name', None),
                'event.action': source.get('event', {}).get('action', None),
                'source.ip': source.get('source', {}).get('ip', None),
                'destination.ip': source.get('destination', {}).get('ip', None),
                'http.request.body.content': source.get('http', {}).get('request', {}).get('body', {}).get('content', None),
                'dns.question.name': source.get('dns', {}).get('question',{}).get('name',None)
            }) 
        return logs

    except ConnectionError as e:
        print(f"ConnectionError connecting to Elasticsearch: {e}")
    except TransportError as e:
        print(f"TransportError connecting to Elasticsearch: {e}")
    except NotFoundError as e:
        print(f"NotFoundError: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    return []

#=====================================================================================================

#========================== GET SIGN IN LOGS FOR USER ====================================

def get_signin_logs(user, timestamp, API_KEY):
    # Connect to Elasticsearch
    try:


        # Initialize the Elasticsearch client
        client = Elasticsearch("https://chromewell-elk-soc.es.us-central1.gcp.cloud.es.io",
        api_key=API_KEY
        )

        
        index_name = ".ds-logs-o365.audit-default*"
        # Define the query with size inside the body
        
        query = f'''
        {{
            "size": 1000,
            "query": {{
                "bool": {{
                    "must": [
                        {{"match": {{"event.dataset": "o365.audit"}}}},
                        {{"terms": {{"event.action": ["UserLoggeIn", "UserLoginFailed"]}}}},
                        {{"range": {{"@timestamp": {{"gte": "now-{timestamp}h", "lte": "now/h"}}}}}},
                        {{"match": {{"user.name":"{user}"}}}}
                    ]
                }}
            }}
        }}
        '''
        
        

        # Fetch logs from Elasticsearch
        response = client.search(index=index_name, body=query)

        # Extract destination IPs from logs
        logs = []
        for hit in response['hits']['hits']:
            source = hit['_source']
            logs.append({
                '@timestamp': source.get('@timestamp', None),
                'event.action': source.get('event', {}).get('action', None),
                'user.name': source.get('user', {}).get('name', None),
                'host.name': source.get('host', {}).get('name', None),
                'source.ip': source.get('source', {}).get('ip', None),
                'source.geo.country_name': source.get('source', {}).get('geo', {}).get('country_name', None)
                
            }) 
        return logs

    except ConnectionError as e:
        print(f"ConnectionError connecting to Elasticsearch: {e}")
    except TransportError as e:
        print(f"TransportError connecting to Elasticsearch: {e}")
    except NotFoundError as e:
        print(f"NotFoundError: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    return []


#-=====================================================================================================================



#=============================== get-file download logs ====================================

def file_download_logs(user, timestamp, API_KEY):
    # Connect to Elasticsearch
    try:


        # Initialize the Elasticsearch client
        client = Elasticsearch("https://chromewell-elk-soc.es.us-central1.gcp.cloud.es.io",
        api_key=API_KEY
        )

        index_name = ".ds-logs-endpoint.events.file-default*"
        # Define the query with size inside the body
        
        query = f'''
        {{
            "size": 1000,
            "query": {{
                "bool": {{
                    "must": [
                        {{"match": {{"event.dataset": "endpoint.events.file"}}}},
                        {{"terms": {{"file.extension": ["exe", "bat", "ps1", "vbs", "xlsx", "pdf", "doc", "pptx", "vbs", "js", "msi", "hta"]}}}},
                        {{"range": {{"@timestamp": {{"gte": "now-{timestamp}h", "lte": "now/h"}}}}}},
                        {{"match": {{"user.name":"{user}"}}}}
                    ]
                }}
            }}
        }}
        '''
        

        # Fetch logs from Elasticsearch
        response = client.search(index=index_name, body=query)

        # Extract destination IPs from logs
        logs = []
        for hit in response['hits']['hits']:
            source = hit['_source']
            logs.append({
                '@timestamp': source.get('@timestamp', None),
                'event.action': source.get('event', {}).get('action', None),
                'user.name': source.get('user', {}).get('name', None),
                'host.name': source.get('host', {}).get('name', None),
                'file.name': source.get('file', {}).get('name', None),
                'file.path': source.get('file', {}).get('path', None),
                'file.extension': source.get('file', {}).get('extension', None),

                
            }) 
        return logs

    except ConnectionError as e:
        print(f"ConnectionError connecting to Elasticsearch: {e}")
    except TransportError as e:
        print(f"TransportError connecting to Elasticsearch: {e}")
    except NotFoundError as e:
        print(f"NotFoundError: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    return []

#===============================================================================



#============================ CAPTURING LOGS =======================================

def capturing_context_logs(in_dir, out_dir, http_query, dns_query, timestamp, API_KEY):

    available_files = ["output_http.xlsx", "output_dns.xlsx"]

    file_path = os.path.join(in_dir, available_files[0])
    today = datetime.now().strftime("%m-%d")

    
    if os.path.exists(file_path):
        try:
            df = pd.read_excel(file_path)
            confirmed_mal = df[df['Type']=="confirmed_malicious"].copy()
            confirmed_mal.drop_duplicates(subset=['user.name'], keep='last', inplace=True)
            
            for index, row in confirmed_mal.iterrows():
                user = row['user.name']

                xls_dir_user = f"{out_dir}/{user}"
                if not os.path.exists(xls_dir_user):
                    os.makedirs(xls_dir_user)

                print(f"[+] Capturing Raw DNS/HTTP Logs for user {user}")
                http_dns_logs = get_compromised_http_dns_logs(user, http_query, dns_query, timestamp, API_KEY)
                
                http_dns_filepath = os.path.join(xls_dir_user, f"http_dns_logs_{user}.xlsx")
                if (http_dns_logs):
                    df_dns = pd.DataFrame(http_dns_logs)
                    df_dns.to_excel(http_dns_filepath, index=False)
                    print(f"[+] HTTP and DNS logs for user {user} have been exported to {http_dns_filepath}]\n")
                else:
                    print(f"[+] No HTTP/DNS logs found for user {user}\n")

                print(f"[+] Capturing Raw Sign-in Logs for user {user}")
                signin_logs = get_signin_logs(user, timestamp, API_KEY)

                signin_logs_filepath = os.path.join(xls_dir_user, f"signin_logs_{user}.xlsx")
                if (signin_logs):
                    df_signin = pd.DataFrame(signin_logs)
                    df_signin.to_excel(signin_logs_filepath, index=False)
                    print(f"[+] Sign-in logs for user {user} have been exported to {signin_logs_filepath}]\n")
                else:
                    print(f"[+] No Sign-In logs found for user {user}\n")

                print(f"[+] Capturing Raw File Logs for user {user}")
                file_logs = file_download_logs(user, timestamp, API_KEY)

                file_download_logs_filepath = os.path.join(xls_dir_user, f"file_download_logs_{user}.xlsx")
                if(file_download_logs):
                    if (file_logs):
                        df_file = pd.DataFrame(file_logs)
                        df_file.to_excel(file_download_logs_filepath, index=False)
                        print(f"[+] File Creation logs for user {user} have been exported to {file_download_logs_filepath}]\n")
                    else:
                        print(f"[+] No File logs found for user {user}\n")

        except Exception as e:
            print(f"An error occurred while reading the file: {e}")

    file_path = os.path.join(in_dir, available_files[1])
    today = datetime.now().strftime("%m-%d")

    
    if os.path.exists(file_path):
        try:
            df = pd.read_excel(file_path)
            confirmed_mal = df[df['Type']=="confirmed_malicious"].copy()
            confirmed_mal.drop_duplicates(subset=['user.name'], keep='last', inplace=True)
            
            for index, row in confirmed_mal.iterrows():
                user = row['user.name']

                xls_dir_user = f"{out_dir}/{user}"
                if not os.path.exists(xls_dir_user):
                    os.makedirs(xls_dir_user)

                print(f"[+] Capturing Raw DNS/HTTP Logs for user {user}")
                http_dns_logs = get_compromised_http_dns_logs(user, http_query, dns_query, timestamp, API_KEY)
                
                http_dns_filepath = os.path.join(xls_dir_user, f"http_dns_logs_{user}.xlsx")
                if (http_dns_logs):
                    df_dns = pd.DataFrame(http_dns_logs)
                    df_dns.to_excel(http_dns_filepath, index=False)
                    print(f"[+] HTTP and DNS logs for user {user} have been exported to {http_dns_filepath}]\n")
                else:
                    print(f"[+] No HTTP/DNS logs found for user {user}\n")

                print(f"[+] Capturing Raw Sign-in Logs for user {user}")
                signin_logs = get_signin_logs(user, timestamp, API_KEY)

                signin_logs_filepath = os.path.join(xls_dir_user, f"signin_logs_{user}.xlsx")
                if (signin_logs):
                    df_signin = pd.DataFrame(signin_logs)
                    df_signin.to_excel(signin_logs_filepath, index=False)
                    print(f"[+] Sign-in logs for user {user} have been exported to {signin_logs_filepath}]\n")
                else:
                    print(f"[+] No Sign-In logs found for user {user}\n")

                print(f"[+] Capturing Raw File Logs for user {user}")
                file_logs = file_download_logs(user, timestamp, API_KEY)

                file_download_logs_filepath = os.path.join(xls_dir_user, f"file_download_logs_{user}.xlsx")
                if(file_download_logs):
                    if (file_logs):
                        df_file = pd.DataFrame(file_logs)
                        df_file.to_excel(file_download_logs_filepath, index=False)
                        print(f"[+] File Creation logs for user {user} have been exported to {file_download_logs_filepath}]\n")
                    else:
                        print(f"[+] No File logs found for user {user}\n")

        except Exception as e:
            print(f"An error occurred while reading the file: {e}")

#============================================================================

# ====================== get endpoint id ==================================

def get_endpoint_id(base_url, api_key, host_name):

    endpoint = f"{base_url}/api/endpoint/metadata"
    
    # Construct the curl command
    curl_command = [
        "curl", "-s", "-XGET", endpoint,
        "-H", "kbn-xsrf: reporting",
        "-H", f"Authorization: ApiKey {api_key}"
    ]

    result = subprocess.run(curl_command, capture_output=True, text=True)
    #print(result)
    
    if result.returncode == 0:
        endpoints_data = json.loads(result.stdout)
        endpoints=endpoints_data["data"]

        for endpoint in endpoints:
            metadata = endpoint.get('metadata', {})
            host_info = metadata.get('host', {})
            hostname = host_info.get('hostname')

            if hostname == host_name:
                agent_id = metadata.get('agent', {}).get('id')
                return agent_id
            else:
                print(f"Host '{host_name}' not found in endpoint metadata.")
                return None
        
    else:
        print(f"Failed to retrieve endpoints. Return Code: {result.returncode}")
        print(f"Response: {result.stderr}")
        return None
    
#==========================================================================

#======================== Isolate Host ====================================

def isolate_host(base_url, api_key, endpoint_id, host_name):
    
    endpoint = f"{base_url}/api/endpoint/action/isolate"
    payload = {
        "endpoint_ids": [endpoint_id]
    }

    # Construct the curl command
    
    curl_command = [
        "curl", "-s", "-XPOST", endpoint,
        "-H", "kbn-xsrf: reporting",
        "-H", "Content-Type: application/json",
        "-H", f"Authorization: ApiKey {api_key}",
        "-d", json.dumps(payload)
    ]

    subprocess.run(curl_command, capture_output=True, text=True)
    
    time.sleep(120)
    
    result_isolate = f"{base_url}/api/endpoint/action"

    curl_command2 = [
        "curl", "-s", "-XGET", result_isolate,
        "-H", "kbn-xsrf: reporting",
        "-H", "Content-Type: application/json",
        "-H", f"Authorization: ApiKey {api_key}"
    ]

    result = subprocess.run(curl_command2, capture_output=True, text=True)
    
    if result.returncode == 0:
        response = json.loads(result.stdout)

        for item in response["data"]:
            command = item.get("command")
            was_successful = item.get("wasSuccessful")

            if command == "isolate" and was_successful:
                isolation_successful = True
                print(f"\nHost {host_name} has been successfully isolated !\n")
                break
        
        if not isolation_successful:
            print(f"Failed to isolate host {host_name}.")

    else:
        print(f"Failed to isolate host {host_name}. Return Code: {result.returncode}")
        print(f"Response: {result.stderr}")

#===========================================================================================

#======================== Initiate Host Isolation =================================

def initiate_host_isolation(in_dir, ELASTIC_BASE_URL, API_KEY):

    available_files = ["output_http.xlsx", "output_dns.xlsx"]

    file_path = os.path.join(in_dir, available_files[0])
    df = pd.read_excel(file_path)

    for index, row in df.iterrows():
        host_name = row['host.name']
        if row['Type'] == "confirmed_malicious":
            endpoint_id = get_endpoint_id(ELASTIC_BASE_URL, API_KEY, host_name)
            if endpoint_id is None:
                print(f"Endpoint ID not found for host {host_name}.")
            else:
                isolate_host(ELASTIC_BASE_URL, API_KEY, endpoint_id, host_name)

    file_path = os.path.join(in_dir, available_files[1])
    df = pd.read_excel(file_path)
    
    for index, row in df.iterrows():
        host_name = row['host.name']
        if row['Type'] == "confirmed_malicious":
            endpoint_id = get_endpoint_id(ELASTIC_BASE_URL, API_KEY, host_name)
            if endpoint_id is None:
                print(f"Endpoint ID not found for host {host_name}.")
            else:
                isolate_host(ELASTIC_BASE_URL, API_KEY, endpoint_id, host_name)

#======================================================================================



#================================= RESPONSE ===========================================


def response(final_results, in_dir, http_query, dns_query, timestamp: str, ELASTIC_BASE_URL, API_KEY):

    today = datetime.now().strftime("%m-%d")
    out_dir = f"C:/Users/prajs_28/OneDrive/Desktop/PKF ALgosmic/Automate IR/Automate-Incident-Response/Compromised_logs_{today}"

    if timestamp.endswith('h'):
        tmstmp = int(timestamp[:-1])

    hours = tmstmp*2
    #isolate_hosts(in_dir, ELASTIC_BASE_URL, API_KEY)
    capturing_context_logs(in_dir, out_dir, http_query, dns_query, hours, API_KEY)
    send_email(final_results, in_dir)
    

#=======================================================================================



#============================= FORMAT EXCEL FILE ============================================

def colourfill(output_excel):    
    df = pd.read_excel(output_excel)

#   Save the DataFrame to an Excel file without formatting first
    df.to_excel(output_excel, index=False, engine='openpyxl')

    # Open the saved Excel file for formatting
    wb = openpyxl.load_workbook(output_excel)
    ws = wb.active

    # Define the colors for the types
    colors = {
        'confirmed_malicious': 'FFC7CE',  # Light red
        'suspicious': 'FFEB9C',          # Light yellow
        'clean': 'C6EFCE'                # Light green
    }

    # Apply the conditional formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[-1]  # Last cell in the row (Type column)
        cell_value = cell.value
        fill_color = colors.get(cell_value, None)
        if fill_color:
            for cell_in_row in row:
                cell_in_row.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# Save the formatted Excel file
    wb.save(output_excel)

#====================================================================================================




 #================================== MAIN ========================================


@app.command()
def main(timestamp: str = typer.Option(..., help="Timestamp range for the query, e.g., '24h', '48h', etc.")):

#------------------------- Declarations -------------------------------------------------------

    final_results_http=[]
    final_results_dns=[]
    final_results=[]
    xls_dir = "C:/Users/prajs_28/OneDrive/Desktop/PKF ALgosmic/Automate IR/Automate-Incident-Response/excel_logs"
    xls_dir_test = "C:/Users/prajs_28/OneDrive/Desktop/PKF ALgosmic/Automate IR/Automate-Incident-Response/test_logs"
    excel_file_path_http = os.path.join(xls_dir, "http_network_logs.xlsx")
    excel_file_path_http_test = os.path.join(xls_dir_test, "http_network_logs.xlsx")
    excel_file_path_dns_test = os.path.join(xls_dir_test, "dns_network_logs.xlsx")
    excel_file_path_dns = os.path.join(xls_dir, "dns_network_logs.xlsx")
    out_dir = "C:/Users/prajs_28/OneDrive/Desktop/PKF ALgosmic/Automate IR/Automate-Incident-Response/CSV Results"
    output_excel_http = "C:/Users/prajs_28/OneDrive/Desktop/PKF ALgosmic/Automate IR/Automate-Incident-Response/CSV Results/output_http.xlsx"
    output_excel_dns = "C:/Users/prajs_28/OneDrive/Desktop/PKF ALgosmic/Automate IR/Automate-Incident-Response/CSV Results/output_dns.xlsx"
    elk_api_path = "C:/Users/prajs_28/OneDrive/Desktop/PKF ALgosmic/Automate IR/Automate-Incident-Response/api_key.txt"
    vt_api_path = "C:/Users/prajs_28/OneDrive/Desktop/PKF ALgosmic/Automate IR/Automate-Incident-Response/vt_api_key.txt"
    ELASTIC_BASE_URL = "https://chromewell-elk-soc.es.us-central1.gcp.cloud.es.io" 
    API_KEY = read_api_key(elk_api_path)

#----------------------------------------------------------------------------------------------

    
#---------------- Elastic Query Builder Code snippet ------------------------------------------
    excel_file_path_fp_domains = "C:/Users/prajs_28/OneDrive/Desktop/PKF ALgosmic/Automate IR/Automate-Incident-Response/false_positive_domains.xlsx"
    # Read domains from Excel
    domains = read_false_positive_domains(excel_file_path_fp_domains)

    # Build Elasticsearch query
    es_query_dns = build_elasticsearch_query_dns(domains)
    es_query_http = build_elasticsearch_query_http(domains)
   
#----------------------------------------------------------------------------------------------


#-------------------------- COLLECT - Retrieve logs from ELK and save as Excel -----------------

    logs_http = get_http_network_logs('.ds-logs-endpoint.events.network-default*', timestamp, es_query_http, API_KEY)
    if logs_http:
        # Convert logs to pandas DataFrame
        df_http = pd.DataFrame(logs_http)
        df_http['dns'] = df_http['http.request.body.content'].apply(extract_dns)
        df_http.drop(columns=['http.request.body.content'], inplace=True)
        df_http.drop_duplicates(subset=['destination.ip'], keep='last', inplace=True)
        df_http.to_excel(excel_file_path_http, index=False)
        print("HTTP logs have been exported to http_network_logs.xls\n\n")
    else:
        print("No logs found for the specified query.")

   
    logs_dns = get_dns_network_logs('.ds-logs-endpoint.events.network-default*',timestamp, es_query_dns, API_KEY)
    if logs_dns:
        # Convert logs to pandas DataFrame
        df_dns = pd.DataFrame(logs_dns)
        df_dns.drop_duplicates(subset=['dns.question.name'], keep='last', inplace=True)
        df_dns.to_excel(excel_file_path_dns, index=False)
        print("DNS logs have been exported to dns_network_logs.xls\n\n")
    else:
        print("No logs found for the specified query.")
    
    '''
#--------------------------------------------------------------------------------------------------
    


    #FOR TESTING PURPOSE..COMMENT OUT THIS SECTION IN ACTUAL PRODUCTION
    #===============================================================================
    df = df.head(5) 
    df.to_excel(excel_file_path_http, index=False)
    print("HTTP Logs have been reduced")
    #===============================================================================


    #FOR TESTING PURPOSE..COMMENT OUT THIS SECTION IN ACTUAL PRODUCTION
    #===============================================================================
    df = df.head(5) 
    df.to_excel(excel_file_path_dns, index=False)
    print("DNS Logs have been reduced")
    #===============================================================================


    '''





#----------------------------- ANALYZE Calling VT and storing results ---------------------------------------
    vt_api_key = get_vt_api_key(vt_api_path)
    vt_client = vt.Client(vt_api_key)


    # For testing purpose change this file path to test log file
    df_http = pd.read_excel(excel_file_path_http_test)

    print("************************** SUMMMARY OF RESULTS FOR HTTP NETWORK CONNECTIONS ***************************\n")
    final_results_http=process_logs_http(vt_client, df_http)
    confirmed_df = pd.DataFrame(final_results_http['malicious_http'])
    suspicious_df = pd.DataFrame(final_results_http['suspicious_http'])
    clean_df = pd.DataFrame(final_results_http['clean_http'])

    # Add 'Type' column
    confirmed_df['Type'] = 'confirmed_malicious'
    suspicious_df['Type'] = 'suspicious'
    clean_df['Type'] = 'clean'

    # Concatenate all results
    results_df = pd.concat([confirmed_df, suspicious_df, clean_df])

    # Merge the original DataFrame with the results DataFrame
    df_updated = pd.merge(df_http, results_df, how='left', left_on='destination.ip', right_on='IP')

    # Drop the redundant 'IP' column from the results
    df_updated = df_updated.drop(columns=['IP'])
    df_updated = df_updated.drop(columns=['User Name'])
    df_updated = df_updated.drop(columns=['Host Name'])
    df_updated = df_updated.drop(columns=['Status'])

    # Defang destination.ip and dns column inline using lambda function
    df_updated['destination.ip'] = df_updated['destination.ip'].apply(lambda ip: ip.replace('.', '[.]'))
    df_updated['dns'] = df_updated['dns'].apply(lambda dns: dns.replace('.', '[.]'))
          
    df_updated.to_excel(output_excel_http, index=False)
    df = pd.read_excel(output_excel_http)

    # Check if the 'Type' column exists in the DataFrame
    if 'Type' in df.columns:
        # Use value_counts to get the count of each type
        type_counts = df['Type'].value_counts()

        for type_name, count in type_counts.items():
            print(f"{type_name}: {count}")

   

    colourfill(output_excel_http)
    print(f"\nView Detailed Results in: '{output_excel_http}'\n")

    
    time.sleep(2)

    # For testing purpose change this file path to test log file
    df_dns = pd.read_excel(excel_file_path_dns_test)

    print("************************** SUMMARY OF RESULTS FOR DNS NETWORK CONNECTIONS ***************************\n")
    final_results_dns=process_logs_dns(vt_client, df_dns)

    confirmed_df = pd.DataFrame(final_results_dns['malicious_dns'])
    suspicious_df = pd.DataFrame(final_results_dns['suspicious_dns'])
    clean_df = pd.DataFrame(final_results_dns['clean_dns'])

    # Add 'Type' column
    confirmed_df['Type'] = 'confirmed_malicious'
    suspicious_df['Type'] = 'suspicious'
    clean_df['Type'] = 'clean'

    # Concatenate all results
    results_df = pd.concat([confirmed_df, suspicious_df, clean_df])

    # Merge the original DataFrame with the results DataFrame
    df_updated = pd.merge(df_dns, results_df, how='left', left_on='dns.question.name', right_on='DNS')

    # Drop the redundant 'IP' column from the results
    df_updated = df_updated.drop(columns=['User Name'])
    df_updated = df_updated.drop(columns=['Host Name'])
    df_updated = df_updated.drop(columns=['Status'])
    df_updated = df_updated.drop(columns=['dns.question.name'])
    df_updated['DNS'] = df_updated['DNS'].apply(lambda dns: dns.replace('.', '[.]'))
    df_updated.to_excel(output_excel_dns, index=False)
    df = pd.read_excel(output_excel_dns)

    # Check if the 'Type' column exists in the DataFrame
    if 'Type' in df.columns:
        # Use value_counts to get the count of each type
        type_counts = df['Type'].value_counts()

        for type_name, count in type_counts.items():
            print(f"{type_name}: {count}")
    
    colourfill(output_excel_dns)
    print(f"\nView Detailed Results in: '{output_excel_dns}'\n")      
    

    final_results = {
        'http': final_results_http,
        'dns': final_results_dns
    }
#---------------------------------------------------------------------------------------------------------


#--------------------------------- RESPOND -------------------------------------------------------------

    print("================================ TAKING RESPONSIVE ACTIONS ============================")
    response(final_results, out_dir, es_query_http, es_query_dns, timestamp, ELASTIC_BASE_URL, API_KEY)

#---------------------------------------------------------------------------------------------------------

    vt_client.close()


if __name__ == "__main__":
    app()








